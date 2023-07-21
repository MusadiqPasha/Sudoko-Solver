import cv2
import numpy as np
import time
import os
import openpyxl
import pyautogui
import pytesseract
from PIL import Image as im
import pyautogui as pg
import easygui
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
time.sleep(2)
print("STARTED")

def takeSS(mode):

    if mode=="USDOKO.COM":
        # USDOKO - region=(265 , 215 , 430 , 430))
        pg.screenshot(r"region.png", region=(265 , 215 , 430 , 430))

    elif mode=="LIVESUDOKO.COM":
        # live sudokko - region=(100 , 270 , 450 ,445)) (zoom web browser 90%)
        # online match livesudoko region=(300 , 270 , 430 ,425))
        pg.screenshot(r"region.png", region=(100 , 270 , 450 ,445))

    else:
        # SUDOKO.COM - region=(25, 250, 350, 350))
        pg.screenshot(r"region.png", region=(25, 250, 350, 350))

    print("screenshot done!")
    pathImage = r"region.png"
    heightImg = 450
    widthImg = 450
    img = cv2.imread(pathImage)
    img = cv2.resize(img, (widthImg, heightImg))

    return img

# used to split the images into 9*9 smol images
def splitBoxes(img):
    rows = np.vsplit(img,9)
    boxes=[]
    for r in rows:
        cols= np.hsplit(r,9)
        for box in cols:
            boxes.append(box)
    return boxes

# used to save all the indv images
def indv_images_make(boxesd):
    i=1
    for image in boxesd:
        data = im.fromarray(image)
        data.save(fr"ALL_NUMBERS\{i}.png")
        i+=1


# used to perform img to text (number)
def getPredection():
    image_folder = r"ALL_NUMBERS"
    detected_numbers = []
    kk = 0
    for _ in os.listdir(image_folder):
        kk += 1
        image_path = fr"{image_folder}\{kk}.png"

        image = cv2.imread(image_path,0)

        # Convert the image to grayscale
        #gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Apply thresholding to create a binary image
        #thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
        thresh = cv2.adaptiveThreshold(image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 31, 30)

        # Perform number detection using OCR
        numbers = pytesseract.image_to_string(thresh, config='--psm 7 --oem 3 -c tessedit_char_whitelist=0123456789')

        # Store the detected numbers in the array

        if numbers == "":
            detected_numbers.append(0)
        else:
            detected_numbers.append(int(numbers))

    return detected_numbers


# used to put those numbers into the website
def write_the_ans(ansz,game):

    if game == "LIVESUDOKO.COM":
        # #  X   , Y
        # # X:  320 Y:  285 RGB: for online math with players
        # # 120 , 285 - LIVESUDOKO (goes right to nxt line) hence
        single_list = [element for sublist in ansz for element in sublist]
        pyautogui.click(120, 285)
        pyautogui.click(120, 285)
        for i in range(81):
            pyautogui.sleep(0.1)
            pyautogui.press(str(single_list[i]))
            pyautogui.sleep(0.1)
            pyautogui.press("right")

    #################################################

    if game == "USDOKO.COM" :
        # # USDOKO
        time.sleep(2)
        single_list = [element for sublist in ansz for element in sublist]
        xi = 270
        yi = 235
        kk=0
        for y in range(9):
            for x in range(9):
                pyautogui.click(xi, yi)
                pyautogui.click(xi, yi)
                pyautogui.sleep(0.1)
                pyautogui.press(str(single_list[kk]))
                pyautogui.press(str(single_list[kk]))
                kk=kk+1
                pyautogui.sleep(0.1)
                xi = xi + 49

            xi = 270
            yi = yi + 48


    # #################################################

    if game == "SUDOKO.COM":
        # # # SUDOKO ONLLINE
        # # 40 , 270 + 53 - SUDOKO
        y = 270
        for row in ansz:
            pyautogui.sleep(0.2)
            pyautogui.click(40, y)
            pyautogui.click(40, y)
            pyautogui.sleep(0.1)
            for i in range(9):
                pyautogui.press(str(row[i]))
                pyautogui.sleep(0.4)
                pyautogui.press('right')
                pyautogui.sleep(0.4)

            y=y+20



# used to display the numbers
def printing(arr):
    for i in range(N):
        for j in range(N):
            print(arr[i][j], end = " ")
        print()

# checking if the block is safe or not
def isSafe(grid, row, col, num):
    for x in range(9):
        if grid[row][x] == num:
            return False

    for x in range(9):
        if grid[x][col] == num:
            return False

    startRow = row - row % 3
    startCol = col - col % 3
    for i in range(3):
        for j in range(3):
            if grid[i + startRow][j + startCol] == num:
                return False
    return True

# solve the sudoko
def solveSudoku(grid, row, col):
    if (row == N - 1 and col == N):
        return True

    if col == N:
        row += 1
        col = 0

    if grid[row][col] > 0:
        return solveSudoku(grid, row, col + 1)
    for num in range(1, N + 1, 1):

            if isSafe(grid, row, col, num):
                grid[row][col] = num

                if solveSudoku(grid, row, col + 1):
                    return True

            grid[row][col] = 0

    return False


###########################--MAIN--###############################################

modee = easygui.buttonbox('Click on which SUDOKO game you are using right now.', 'CHOICE', ('LIVESUDOKO.COM', 'USDOKO.COM', 'SUDOKO.COM'))
img = takeSS(modee)
boxesx= splitBoxes(img) # split into boexs of array
indv_images_make(boxesx)  # download em arrays
final_numbers = getPredection()    # find the numbers from image

two_d_list = [final_numbers[i:i+9] for i in range(0, 81, 9)] # 1D TO 2D LIST
workbook = openpyxl.Workbook()
sheet = workbook.active
# write em to the excel file to show em
for row_index, row in enumerate(two_d_list, start=1):
    for col_index, value in enumerate(row, start=1):
        sheet.cell(row=row_index, column=col_index).value = value

workbook.save(r"output.xlsx")
print("###############################-GIVEN-#####################################")
os.startfile(r"output.xlsx")
pyautogui.alert("Check if the detected numbers are correct.\n  If not then change them in the excel file.\n         SAVE the file before closing it.")


result = pyautogui.confirm(text='Are all the numbers now correct?', title='Confirmation', buttons=['YES', 'NO'])

if result == 'YES':
    print("Processing the data...")

    workbook = openpyxl.load_workbook(r'output.xlsx')
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        row_data = [int(cell) for cell in row if isinstance(cell, (int, float))]
        data.append(row_data)
    print("###############################-GIVEN FINAL-#####################################")
    for row in data:
        print(row)

else:
    print("Data processing canceled.")


print("#################################-NOW SOLVING-########################################")

N = 9
solveSudoku(data, 0, 0)
printing(data)
time.sleep(2)

print("#################################-NOW TYPING!!-########################################")

write_the_ans(data,modee)
